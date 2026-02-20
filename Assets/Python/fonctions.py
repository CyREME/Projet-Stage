import sys
import os
import json
import base64
import pandas as pd
from openpyxl import load_workbook
from Crypto.Cipher import AES
from Crypto.Util import Counter

# --- CONFIGURATION SÉCURITÉ ---
ENCRYPTION_KEY = "bV9zYjYmNypAIVpLdzJRM3U1eTh4eiYh"

class EncryptionService:
    @staticmethod
    def encrypt(data):
        if not data: data = ""
        key = ENCRYPTION_KEY.encode('utf-8')
        iv = os.urandom(16)
        ctr = Counter.new(128, initial_value=int.from_bytes(iv, byteorder='big'))
        cipher = AES.new(key, AES.MODE_CTR, counter=ctr)
        encrypted = cipher.encrypt(data.encode('utf-8'))
        combined = iv + encrypted
        return base64.b64encode(combined).decode('utf-8')

    @staticmethod
    def decrypt(encrypted_data):
        if not encrypted_data: return ""
        try:
            raw_data = base64.b64decode(encrypted_data)
            iv = raw_data[:16]
            payload = raw_data[16:]
            key = ENCRYPTION_KEY.encode('utf-8')
            ctr = Counter.new(128, initial_value=int.from_bytes(iv, byteorder='big'))
            cipher = AES.new(key, AES.MODE_CTR, counter=ctr)
            return cipher.decrypt(payload).decode('utf-8')
        except:
            return ""

# --- LOGIQUE DE FICHIERS ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FILE_PATH = os.path.join(BASE_DIR, '..', 'Temp', 'import.xlsx')
JSON_PATH = os.path.join(BASE_DIR, '..', 'Json', 'contacts.json')

def check_Json():
    if os.path.exists(JSON_PATH):
        with open(JSON_PATH, 'r', encoding='utf-8') as file:
            try:
                return json.load(file)
            except json.JSONDecodeError:
                return {}
    return {}

def contact_existe_deja(nom_neuf, prenom_neuf, dict_contacts, encryptor):
    for c in dict_contacts.values():
        # Déchiffrement des données du JSON pour la comparaison
        nom_existant = encryptor.decrypt(c.get("nom", ""))
        prenom_existant = encryptor.decrypt(c.get("prenom", ""))

        if nom_existant.upper() == nom_neuf.upper() and prenom_existant.lower() == prenom_neuf.lower():
            return True
    return False

def merged_Cells():
    wb = load_workbook(FILE_PATH)
    ws = wb.active
    for merged_range in list(ws.merged_cells.ranges):
        borne = str(merged_range)
        val = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
        ws.unmerge_cells(borne)
        for row in ws[borne]:
            for cell in row:
                cell.value = val
    wb.save(FILE_PATH)
    wb.close()

def get_data():
    merged_Cells()
    df = pd.read_excel(FILE_PATH, dtype="str")
    dictContacts = check_Json()

    # Trouver le prochain ID disponible
    ids = [int(k) for k in dictContacts.keys() if k.isdigit()]
    indexContact = max(ids) + 1 if ids else 1

    encryptor = EncryptionService()

    for index, row in df.iterrows():
        # 1. Extraction et nettoyage Nom/Prénom
        nom_complet = str(row.get('NOMS', '')).split()
        nom_brut = " ".join([w for w in nom_complet if w.isupper()]).strip()
        prenom_brut = " ".join([w for w in nom_complet if not w.isupper()]).strip()

        # 2. VÉRIFICATION DES DOUBLONS (Compare le clair du Excel avec le déchiffré du JSON)
        if contact_existe_deja(nom_brut, prenom_brut, dictContacts, encryptor):
            continue # Si existe, on passe à la ligne suivante

        # 3. Préparation des autres données
        data_row = {
            "nom": nom_brut,
            "prenom": prenom_brut,
            "service": str(row.get('SERVICE', '')).replace('nan', '').strip(),
            "fonctions": str(row.get('FONCTIONS', '')).replace('nan', '').strip(),
            "numInterne": str(row.get('Interne', '')).replace('nan', '').strip(),
            "numMobile": str(row.get('TEL MOBILE', '')).replace('nan', '').strip(),
            "numFixe": str(row.get('TEL FIXE', '')).replace('nan', '').strip()
        }

        # Formatage des numéros
        for key in ["numMobile", "numFixe"]:
            val = data_row[key].replace("_x000D_", " ")
            if len(val) == 9: val = "0" + val
            data_row[key] = val

        # 4. CHIFFREMENT ET AJOUT
        dictContacts[str(indexContact)] = {
            "id": str(indexContact),
            "nom": encryptor.encrypt(data_row["nom"]),
            "prenom": encryptor.encrypt(data_row["prenom"]),
            "service": encryptor.encrypt(data_row["service"]),
            "fonctions": encryptor.encrypt(data_row["fonctions"]),
            "numInterne": encryptor.encrypt(data_row["numInterne"]),
            "numMobile": encryptor.encrypt(data_row["numMobile"]),
            "numFixe": encryptor.encrypt(data_row["numFixe"])
        }
        indexContact += 1

    # 5. Sauvegarde
    with open(JSON_PATH, 'w', encoding='utf-8') as file:
        json.dump(dictContacts, file, indent=4, ensure_ascii=False)

    if os.path.exists(FILE_PATH):
        os.remove(FILE_PATH)

if __name__ == "__main__":
    try:
        if len(sys.argv) > 1:
            FILE_PATH = sys.argv[1]
        get_data()
        print("SUCCESS")
    except Exception as e:
        print(f"ERREUR PYTHON : {e}")